import os
import csv
import datetime
import pandas as pd
import xlrd
from openpyxl import Workbook
from openpyxl import load_workbook
# xlrd.open_workbook(r'D:/ACM_Code_Lib/other/爬虫/hello1.xlsx')
dfs = []
for fname in os.listdir("./"):
    # print(fname)
    if fname.endswith(".xlsx") and fname != "hello.xlsx":
        df = pd.read_excel(fname, header = None, sheet_name = None)
        dfs.extend(df.values())
        print(df.keys())
print(dfs)
result = pd.concat(dfs)
# result.to_excel("./hello.xlsx", index = False)
print('------------')
wb = Workbook()
ws = wb.active

ws['A1']=42
ws.append([1,2,3])
ws['A3'] = datetime.datetime.now()

wb.save('sample.xlsx')

wb2 = load_workbook(r"D:\ACM_Code_Lib\other\爬虫\hello.xlsx")

print(wb2.sheetnames)


'''
https://www.zhihu.com/question/35904647
pandas
https://pandas.pydata.org/pandas-docs/version/0.22.0/10min.html
openpyxl
https://www.zhihu.com/question/35904647/answer/1580520424
https://www.jianshu.com/p/90569d5044be
xlrd
https://www.cnblogs.com/shaosks/p/6098282.html
xlrd更新到了2.0.1版本，只支持.xls文件
pip uninstall xlrd
pip install xlrd==1.2.0
也可以用openpyxl代替xlrd打开.xlsx文件:
df=pandas.read_excel('data.xlsx',engine='openpyxl')

with open('jdData.csv', 'a+', newline='', encoding='gb18030') as f:
    writer = csv.writer(f)
    writer.writerow(('留言时间', '评分', '回复数', '点赞数', '图片数', '评论内容'))
'''